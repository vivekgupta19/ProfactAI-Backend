import os
from io import BytesIO
from pathlib import Path
from typing import Dict, List, BinaryIO

import openpyxl
import pandas as pd

from config import EXCEL_DATA_DIR
from . import llm, vectorstore
from database import db


def _iter_document_files(base_dir: str) -> List[Path]:
    """Iterate through all supported document files in a directory"""
    base = Path(base_dir)
    files: List[Path] = []
    supported_extensions = {".xlsx", ".xls", ".csv"}
    
    for entry in base.iterdir():
        if entry.is_file() and entry.suffix.lower() in supported_extensions:
            files.append(entry)
    return files


def _row_to_text(file_name: str, sheet_name: str, row_index: int, row: pd.Series) -> str:
    """Convert a row to searchable text"""
    parts = [f"File: {file_name}", f"Sheet: {sheet_name}"]
    for col, val in row.items():
        if pd.isna(val):
            continue
        text_val = str(val).strip()
        if not text_val:
            continue
        parts.append(f"{col}: {text_val}")
    return " | ".join(parts)


def _read_excel_sheets_from_bytes(data: bytes, file_name: str) -> Dict[str, pd.DataFrame]:
    """
    Robustly read all sheets from an Excel file into a dict of DataFrames.

    We try, in order:
      1) pandas.read_excel (all sheets, default engine)
      2) pandas.read_excel with explicit openpyxl engine
      3) openpyxl.load_workbook + manual conversion to DataFrames

    This is specifically to avoid brittle "tuple index out of range" errors that
    sometimes surface from the underlying Excel engines.
    """

    last_exc: Exception | None = None

    # 1) pandas.read_excel with default engine
    try:
        with BytesIO(data) as bio:
            sheets = pd.read_excel(bio, sheet_name=None, dtype=str)
        cleaned: Dict[str, pd.DataFrame] = {}
        for name, df in sheets.items():
            if isinstance(df, pd.DataFrame) and not df.empty:
                cleaned[name] = df.fillna("")
        if cleaned:
            return cleaned
    except Exception as exc:  # pragma: no cover - defensive
        last_exc = exc

    # 2) pandas.read_excel with explicit openpyxl engine
    try:
        with BytesIO(data) as bio:
            sheets = pd.read_excel(bio, sheet_name=None, dtype=str, engine="openpyxl")
        cleaned = {}
        for name, df in sheets.items():
            if isinstance(df, pd.DataFrame) and not df.empty:
                cleaned[name] = df.fillna("")
        if cleaned:
            return cleaned
    except Exception as exc:  # pragma: no cover - defensive
        last_exc = exc

    # 3) Fallback: use openpyxl directly and build DataFrames sheet-by-sheet
    try:
        wb = openpyxl.load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
        cleaned: Dict[str, pd.DataFrame] = {}
        for ws in wb.worksheets:
            values_iter = ws.iter_rows(values_only=True)
            try:
                headers = next(values_iter)
            except StopIteration:
                continue

            headers = [str(h) if h is not None else "" for h in headers]
            rows = list(values_iter)
            if not rows:
                continue

            df = pd.DataFrame(rows, columns=headers, dtype=str)
            cleaned[ws.title] = df.fillna("")

        if cleaned:
            return cleaned
    except Exception as exc:  # pragma: no cover - defensive
        last_exc = exc

    raise RuntimeError(f"Failed to read Excel file '{file_name}': {last_exc}")


def _index_sheets(
    file_name: str,
    sheets: Dict[str, pd.DataFrame],
    organization_id: str,
    user_id: str
) -> Dict:
    """Index all sheets/rows from a single file into Postgres (pgvector)"""
    total_rows = 0
    added_docs = 0

    batch_texts: List[str] = []
    batch_metas: List[dict] = []
    batch_ids: List[str] = []
    batch_size = 100

    for sheet_name, df in sheets.items():
        if df is None or df.empty:
            continue

        for idx, row in df.iterrows():
            total_rows += 1
            doc_id = f"{organization_id}::{file_name}::{sheet_name}::{idx}"
            text = _row_to_text(file_name, sheet_name, int(idx), row)
            metadata = {
                "file": file_name,
                "sheet": sheet_name,
                "row_index": int(idx),
                "num_columns": len(df.columns),
                "organization_id": organization_id,
                "user_id": user_id,
            }

            batch_texts.append(text)
            batch_metas.append(metadata)
            batch_ids.append(doc_id)

            if len(batch_texts) >= batch_size:
                embeddings = llm.embed_texts(batch_texts)
                vectorstore.add_documents(
                    organization_id=organization_id,
                    texts=batch_texts,
                    metadatas=batch_metas,
                    ids=batch_ids,
                    embeddings=embeddings,
                )
                added_docs += len(batch_texts)
                batch_texts, batch_metas, batch_ids = [], [], []

    # Flush remaining
    if batch_texts:
        embeddings = llm.embed_texts(batch_texts)
        vectorstore.add_documents(
            organization_id=organization_id,
            texts=batch_texts,
            metadatas=batch_metas,
            ids=batch_ids,
            embeddings=embeddings,
        )
        added_docs += len(batch_texts)

    return {
        "fileName": file_name,
        "totalRowsSeen": total_rows,
        "documentsAdded": added_docs,
    }


def build_index(organization_id: str, user_id: str, rebuild: bool = True) -> Dict:
    """
    Read all document files on disk for an organization,
    convert rows to text, embed, and store in Postgres (pgvector).
    """
    if rebuild:
        collection = vectorstore.reset_user_collection(organization_id)
    else:
        collection = vectorstore.get_user_collection(organization_id)

    document_files = _iter_document_files(EXCEL_DATA_DIR)

    agg_total_rows = 0
    agg_added_docs = 0
    per_file_stats = []

    for file_path in document_files:
        file_name = file_path.name
        file_ext = file_path.suffix.lower()
        
        try:
            if file_ext == ".csv":
                # For CSV, treat as single sheet
                df = pd.read_csv(file_path, dtype=str)
                sheets = {"Sheet1": df.fillna("")}
            else:
                with file_path.open("rb") as f:
                    data = f.read()
                sheets = _read_excel_sheets_from_bytes(data, file_name)
        except Exception:
            # Skip unreadable files but keep the job going
            continue

        stats = _index_sheets(file_name, sheets, organization_id, user_id)
        agg_total_rows += stats["totalRowsSeen"]
        agg_added_docs += stats["documentsAdded"]
        per_file_stats.append(stats)
        
        # Record document in database
        try:
            db.create_document(
                collection_id=None,  # We can create a default collection later
                organization_id=organization_id,
                user_id=user_id,
                file_name=file_name,
                file_type=file_ext,
                file_size=file_path.stat().st_size,
                total_rows=stats["totalRowsSeen"],
                indexed_rows=stats["documentsAdded"]
            )
        except:
            pass  # Don't fail if DB insert fails

    return {
        "filesIndexed": [str(p) for p in document_files],
        "perFileStats": per_file_stats,
        "totalRowsSeen": agg_total_rows,
        "documentsAdded": agg_added_docs,
        "collectionName": collection.name,
    }


def index_uploaded_file(
    file_stream: BinaryIO,
    file_name: str,
    organization_id: str,
    user_id: str
) -> Dict:
    """
    Index a single uploaded file (provided as a binary stream) into Postgres (pgvector).
    Does NOT reset the collection. New rows are appended to existing index.
    """
    file_ext = Path(file_name).suffix.lower()
    
    try:
        if file_ext == ".csv":
            df = pd.read_csv(file_stream, dtype=str)
            sheets = {"Sheet1": df.fillna("")}
        else:
            # Read the entire stream into memory once so we can safely retry
            data = file_stream.read()
            sheets = _read_excel_sheets_from_bytes(data, file_name)
    except Exception as exc:
        # Normalize low-level Excel engine errors into a clear message
        raise RuntimeError(f"Failed to read uploaded file '{file_name}': {exc}")

    raw_stats = _index_sheets(file_name, sheets, organization_id, user_id)
    
    # Record document in database
    try:
        db.create_document(
            collection_id=None,
            organization_id=organization_id,
            user_id=user_id,
            file_name=file_name,
            file_type=file_ext,
            file_size=0,  # Size not available from stream
            total_rows=raw_stats["totalRowsSeen"],
            indexed_rows=raw_stats["documentsAdded"]
        )
    except Exception as e:
        print(f"Warning: Could not record document in DB: {e}")

    # Shape stats to what the frontend expects (snake_case keys)
    api_stats = {
        "file_name": raw_stats["fileName"],
        "total_rows_seen": raw_stats["totalRowsSeen"],
        "documents_added": raw_stats["documentsAdded"],
        "collection_name": vectorstore.get_user_collection(organization_id).name,
    }
    return api_stats